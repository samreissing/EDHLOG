#!/usr/bin/env python3
"""Import game log data from the Google Sheets xlsx export into data/seed.json."""

import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT_PATHS = [ROOT / "data" / "seed.json", ROOT / "public" / "data" / "seed.json"]


def normalize_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    m = re.match(r"^(\d{1,2})/\?/(\d{2})$", text)
    if m:
        month, yy = m.groups()
        return f"20{yy}-{month.zfill(2)}-15"
    return text


def clean_deck_name(name: str) -> str:
    return re.sub(r"\s*\*!!!\*\(Retired Deck\)\*!!!\*", "", name).strip()


def get_colors(row) -> list[str]:
    colors = []
    for idx, c in enumerate(["W", "U", "B", "R", "G"], start=5):
        if row[idx] == c:
            colors.append(c)
    return colors


def main(xlsx_path: str) -> None:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Main Sheet"]
    decks: dict[str, dict] = {}
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    for i, row in enumerate(rows):
        if row[11] == "Active Decks":
            j = i + 1
            while j < len(rows) and rows[j][11] and rows[j][11] != "Retired Decks":
                r = rows[j]
                if r[10] is not None and r[11]:
                    name = clean_deck_name(str(r[11]))
                    decks[name] = {
                        "name": name,
                        "bracket": int(r[10]),
                        "colors": get_colors(r),
                        "retired": False,
                    }
                j += 1
        if row[11] == "Retired Decks":
            j = i + 1
            while j < len(rows) and rows[j][10] is not None and rows[j][11]:
                r = rows[j]
                name = clean_deck_name(str(r[11]))
                decks[name] = {
                    "name": name,
                    "bracket": int(r[10]),
                    "colors": get_colors(r),
                    "retired": True,
                }
                j += 1

    games = []
    for year in ["2024", "2025", "2026"]:
        sheet = wb[f"Game Log {year}"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2]:
                d = normalize_date(row[0])
                if len(d) >= 4 and d[:4] != year:
                    d = f"{year}{d[4:]}"
                raw = str(row[1])
                name = clean_deck_name(raw)
                if name not in decks:
                    decks[name] = {
                        "name": name,
                        "bracket": 4,
                        "colors": [],
                        "retired": "Retired" in raw,
                    }
                games.append({"date": d, "deck": name, "result": row[2]})

    first_game_by_deck: dict[str, str] = {}
    for game in games:
        existing = first_game_by_deck.get(game["deck"])
        if not existing or game["date"] < existing:
            first_game_by_deck[game["deck"]] = game["date"]

    deck_list = list(decks.values())
    for deck in deck_list:
        deck["createdAt"] = first_game_by_deck.get(deck["name"], "2024-04-15")

    for i, g in enumerate(games):
        g["id"] = f"game-{i + 1}"

    payload = json.dumps({"decks": deck_list, "games": games}, indent=2)
    for out in OUT_PATHS:
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(payload)
    print(f"Wrote {len(games)} games and {len(decks)} decks to {OUT_PATHS[0]}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python scripts/import_xlsx.py path/to/file.xlsx")
        sys.exit(1)
    main(sys.argv[1])
